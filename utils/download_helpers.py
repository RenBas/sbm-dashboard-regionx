"""Download helpers for generating data collection template and reports."""

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill


def generate_excel_template():
    """
    Create an Excel template with the exact column headers required for the SBM data upload.
    Includes one sample row for guidance (users can delete it).
    """
    # Column headers as used in the actual upload file
    columns = [
        "Region",
        "School ID",
        "School Name",
        "Division",
        "Offering",
        "Latitude",
        "Longitude",
        "CT_1. Grade 3 learners achieve the proficiency level for each cluster of early language, literacy, and numeracy skills",
        "CT_2. Grade 6, 10, and 12 learners achieve the proficiency level in all 21st century skills and core learning areas in the National Achievement Test (NAT)",
        "CT_3. School-based ALS learners attain certification as elementary and junior high school completers",
        "CT_4. Teachers prepare contextualized learning materials responsive to the needs of learners",
        "CT_5. Teachers conduct remediation activities to address learning gaps in reading and comprehension, science and technology, and mathematics",
        "CT_6. Teachers integrate topics promoting peace and DepEd core values",
        "CT_7. The school conducts test item analysis to inform its teaching and learning process",
        "CT_8. The school engages local industries to strengthen its TLE-TVL course offerings",
        "LE_9. The school has zero bullying incidence",
        "LE_10. The school has zero child abuse incidence",
        "LE_11. The school has reduced its drop-out incidence",
        "LE_12. The school conducts culture-sensitive activities",
        "LE_13. The school provides access to learning experiences for the disadvantaged, OSYs, and adult learners",
        "LE_14. The school has a functional school-based ALS program",
        "LE_15. The school has a functional child-protection committee",
        "LE_16. The school has a functional DRRM plan",
        "LE_17. The school has a functional support mechanism for mental wellness",
        "LE_18. The school has special education- and FWD-friendly facilities",
        "LG_19. The school develops a strategic plan",
        "LG_20. The school has a functional school-community planning team",
        "LG_21. The school has a functional Supreme Student Government/ Supreme Pupil Government",
        "LG_22. The school innovates in its provision of frontline services to stakeholders",
        "AC_23. The school's strategic plan is operationalized through an implementation plan",
        "AC_24. The school has a functional School Governance Council (SGC)",
        "AC_25. The school has a functional Parent-Teacher Association (PTA)",
        "AC_26. The school collaborates with stakeholders and other schools in strengthening partnerships",
        "AC_27. The school monitors and evaluates its programs, projects, and activities",
        "AC_28. The school maintains an average rating of satisfactory from its internal and external stakeholders",
        "HR_29. School personnel achieve an average rating of very satisfactory in the individual performance commitment and review",
        "HR_30. The school achieves an average rating of very satisfactory in the office performance commitment and review",
        "HR_31. The school conducts needs-based Learning Action Cells and Learning & Development activities",
        "HR_32. The school facilitates the promotion and continuous professional development of its personnel",
        "HR_33. The school recognizes and rewards milestone achievements of its personnel",
        "HR_34. The school facilitates receipt of correct salaries, allowances, and other additional compensation in a timely manner",
        "HR_35. Teacher workload is distributed fairly and equitably",
        "FR_36. The school inspects its infrastructure and facilities",
        "FR_37. The school initiates improvement of its infrastructure and facilities",
        "FR_38. The school has a functional library",
        "FR_39. The school has functional water, electric, and internet facilities",
        "FR_40. The school has a functional computer laboratory/ classroom",
        "FR_41. The school achieves a 75-100% utilization rate of its Maintenance and Other Operating Expenses (MOOE)",
        "FR_42. The school liquidates 100% of its utilized MOOE"
    ]

    # Create DataFrame with columns only
    df = pd.DataFrame(columns=columns)

    # Add one sample row (optional, can be deleted by user)
    sample_row = {col: "" for col in columns}
    sample_row["Region"] = "Region X"
    sample_row["School ID"] = "123456"
    sample_row["School Name"] = "Sample Elementary School"
    sample_row["Division"] = "Bukidnon"
    sample_row["Offering"] = "Elementary"
    sample_row["Latitude"] = 8.1234
    sample_row["Longitude"] = 124.5678
    # For indicator columns, leave blank or put sample scores (e.g., 0-3)
    # We'll leave them blank so they can fill in; but you could put a few sample numbers
    # Uncomment the next lines if you want sample scores:
    # for i, col in enumerate(columns[7:], start=1):
    #     sample_row[col] = round(random.uniform(0, 3), 1)

    df = pd.concat([df, pd.DataFrame([sample_row])], ignore_index=True)

    # Write to Excel with some formatting
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="SBM Data")
        workbook = writer.book
        worksheet = writer.sheets["SBM Data"]

        # Format header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0033A0", end_color="0033A0", fill_type="solid")
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Set column widths (wider for long indicator descriptions)
        for col_idx, col_name in enumerate(columns, start=1):
            if col_idx <= 7:  # metadata columns
                worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = 18
            else:
                worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = 35

        # Freeze the header row
        worksheet.freeze_panes = 'A2'

    output.seek(0)
    return output


def generate_report_data(division_name, schools_in_sdo, complete_schools):
    """
    Generate a simple CSV report for the division.
    (This is the existing CSV report function, unchanged.)
    """
    if not complete_schools:
        return pd.DataFrame()
    rows = []
    for school in schools_in_sdo:
        rows.append({
            "School": school.get("name", ""),
            "Type": school.get("type", ""),
            "Enrollment": school.get("enrollment", 0),
            "Overall Index": school.get("overall_index", 0),
            **{dim: school.get("dimension_scores", [0]*6)[i] for i, dim in enumerate(DIMENSION_NAMES)},
            "Data Status": school.get("data_status", "Pending")
        })
    return pd.DataFrame(rows)
